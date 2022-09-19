from doctest import master
import settings
import tools
import data_types
import openpyxl as xl



def doTest():
    masterFileWB = xl.load_workbook(settings.workspace.getMasterFileName())

    firstSheet = masterFileWB.sheetnames[0]
    isOk = True
    for col in range(0, 25):
        cellCode = tools.getCellCode(0, col)
        print(f"Col {chr(ord('A') + col)}, val: {masterFileWB[firstSheet][cellCode].value}")
        if len(data_types.masterRecordColumns) <= col:
            if masterFileWB[firstSheet][cellCode].value is not None:
                isOk = False
                print(f"Missing fields in source: {masterFileWB[firstSheet][cellCode].value}")
        elif data_types.masterRecordColumns[col][1][1] != masterFileWB[firstSheet][cellCode].value:
            print(f"Error in source: {data_types.masterRecordColumns[col][1][1]}")
            isOk = False
        for sheet in masterFileWB.sheetnames:
            if masterFileWB[firstSheet][cellCode].value != masterFileWB[sheet][cellCode].value:
                isOk = False
                print (f"{masterFileWB[sheet].title} sheet error")

    if isOk == True:
        print('Data is good')
    else:
        print('Data is NOT good')
    return
    cell = masterFileWB['Frial']['K3']

    for sheet in masterFileWB:
        for col in range(0, 20):
            print(f"Title: {sheet.title}, Column: {chr(ord('A') + col)}, is empty: {tools.isColumnEmpty(sheet, col, 1)}")
            



for c in range(ord('A'), ord('F') + 1):
    print(chr(c))

settings.load()

doTest()